import openpyxl
from datetime import datetime


class XLSX:
    @staticmethod
    def createspreadsheet(menunumber, reqjson):

        # Nome do Workbook
        date = str(datetime.now())
        datename = date[:10]
        name = str(datename) + ".xlsx"

        # Cria ou abre Workbook para registro
        try:
            workbook = openpyxl.load_workbook(name)
        except:
            workbook = openpyxl.Workbook()

        # Planilha de registro dentro do Workbook
        sheet = workbook.worksheets[0]

        # Célula de início de registro das pesquisas
        if sheet.max_row == 1:
            rowhearder = 1
            rowsearch = 2
            dateinfo = "A" + str(rowhearder)
            hourinfo = "B" + str(rowhearder)
            returninfo = "A" + str(rowsearch)
            returninfoterm = "B" + str(rowsearch)
        else:
            rowhearder = sheet.max_row + 2
            rowsearch = sheet.max_row + 3
            dateinfo = "A" + str(rowhearder)
            hourinfo = "B" + str(rowhearder)
            returninfo = "A" + str(rowsearch)
            returninfoterm = "B" + str(rowsearch)

        # Data da Pesquisa é a primeira linha registrada
        sheet[dateinfo] = str(date[:10])
        sheet[hourinfo] = str(date[11:13] + "h" + date[14:16] + "min")
        sheetcellcount = 1

        # Pesquisa por país ou termo
        if menunumber == 1:

            # Tipo de resposta
            sheet[returninfo] = "Pesquisa por pais ou termo:"
            if reqjson["parameters"] == []:
                sheet[returninfoterm] = "Paises Listados na API."
            else:
                sheet[returninfoterm] = "Pesquisa por termo: " + reqjson["parameters"]["search"]

            # Exportar dados
            for countries in reqjson["response"]:
                sheet["A" + str(int(rowsearch) + int(sheetcellcount))] = str(countries)
                sheetcellcount += 1

        # Pesquisa por dadps atuais
        elif menunumber == 2:

            # Tipo de resposta
            sheet[returninfo] = "Pesquisa por dados atuais."
            if reqjson["results"] == 1:
                sheet[returninfoterm] = "Pesquisa por país específico."
            else:
                sheet[returninfoterm] = "Dados de todos os paises."

            # Cabeçalho da planilha
            sheet["A" + str(int(rowsearch) + int(sheetcellcount))] = "País"
            sheet["B" + str(int(rowsearch) + int(sheetcellcount))] = "Total de Casos"
            sheet["C" + str(int(rowsearch) + int(sheetcellcount))] = "Casos Ativos"
            sheet["D" + str(int(rowsearch) + int(sheetcellcount))] = "Casos Críticos"
            sheet["E" + str(int(rowsearch) + int(sheetcellcount))] = "Casos Recuperados"
            sheet["F" + str(int(rowsearch) + int(sheetcellcount))] = "Casos Novos"
            sheet["G" + str(int(rowsearch) + int(sheetcellcount))] = "Casos por Milhão"
            sheet["H" + str(int(rowsearch) + int(sheetcellcount))] = "Total de Mortos"
            sheet["I" + str(int(rowsearch) + int(sheetcellcount))] = "Novas Mortes"
            sheet["J" + str(int(rowsearch) + int(sheetcellcount))] = "Mortes por Milhão"
            sheet["K" + str(int(rowsearch) + int(sheetcellcount))] = "Testes Realizados"
            sheet["L" + str(int(rowsearch) + int(sheetcellcount))] = "Testes por Milhão"
            sheetcellcount += 1

            # Exportar dados
            for response in reqjson["response"]:
                sheet["A" + str(int(rowsearch) + int(sheetcellcount))] = response["country"]
                sheet["B" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["total"]
                sheet["C" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["active"]
                sheet["D" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["critical"]
                sheet["E" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["recovered"]
                if response["cases"]["new"] is None:
                    sheet["F" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["new"]
                elif response["cases"]["new"][0:1] == "+":
                    sheet["F" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["new"][1:]
                else:
                    sheet["F" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["new"]
                sheet["G" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["1M_pop"]
                sheet["H" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["total"]
                if response["deaths"]["new"] is None:
                    sheet["I" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["new"]
                elif response["deaths"]["new"][0:1] == "+":
                    sheet["I" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["new"][1:]
                else:
                    sheet["I" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["new"]
                sheet["J" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["1M_pop"]
                sheet["K" + str(int(rowsearch) + int(sheetcellcount))] = response["tests"]["total"]
                sheet["L" + str(int(rowsearch) + int(sheetcellcount))] = response["tests"]["1M_pop"]
                sheetcellcount += 1

        # Pesquisa por dados no histórico
        else:

            # Tipo de resposta
            sheet[returninfo] = "Pesquisa por dados no histórico."
            if reqjson["parameters"]["country"] == "All":
                sheet[returninfoterm] = "Pesquisa por dados de todos os paises"
            else:
                sheet[returninfoterm] = "Dados de país específico: " + str(reqjson["parameters"]["country"])

            # Cabeçalho da planilha
            sheet["A" + str(int(rowsearch) + int(sheetcellcount))] = "País"
            sheet["B" + str(int(rowsearch) + int(sheetcellcount))] = "Data de Referência"
            sheet["C" + str(int(rowsearch) + int(sheetcellcount))] = "Hora de Referência"
            sheet["D" + str(int(rowsearch) + int(sheetcellcount))] = "Total de Casos"
            sheet["E" + str(int(rowsearch) + int(sheetcellcount))] = "Casos Ativos"
            sheet["F" + str(int(rowsearch) + int(sheetcellcount))] = "Casos Críticos"
            sheet["G" + str(int(rowsearch) + int(sheetcellcount))] = "Casos Recuperados"
            sheet["H" + str(int(rowsearch) + int(sheetcellcount))] = "Casos Novos"
            sheet["I" + str(int(rowsearch) + int(sheetcellcount))] = "Casos por Milhão"
            sheet["J" + str(int(rowsearch) + int(sheetcellcount))] = "Total de Mortos"
            sheet["K" + str(int(rowsearch) + int(sheetcellcount))] = "Novas Mortes"
            sheet["L" + str(int(rowsearch) + int(sheetcellcount))] = "Mortes por Milhão"
            sheet["M" + str(int(rowsearch) + int(sheetcellcount))] = "Testes Realizados"
            sheet["N" + str(int(rowsearch) + int(sheetcellcount))] = "Testes por Milhão"
            sheetcellcount += 1

            # Exportar dados
            for response in reqjson["response"]:
                sheet["A" + str(int(rowsearch) + int(sheetcellcount))] = response["country"]
                sheet["B" + str(int(rowsearch) + int(sheetcellcount))] = response["day"]
                sheet["C" + str(int(rowsearch) + int(sheetcellcount))] = response["time"][11:16]
                sheet["D" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["total"]
                sheet["E" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["active"]
                sheet["F" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["critical"]
                sheet["G" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["recovered"]
                if response["cases"]["new"] is None:
                    sheet["H" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["new"]
                elif response["cases"]["new"][0:1] == "+":
                    sheet["H" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["new"][1:]
                else:
                    sheet["H" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["new"]
                sheet["I" + str(int(rowsearch) + int(sheetcellcount))] = response["cases"]["1M_pop"]
                sheet["J" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["total"]
                if response["deaths"]["new"] is None:
                    sheet["K" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["new"]
                elif response["deaths"]["new"][0:1] == "+":
                    sheet["K" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["new"][1:]
                else:
                    sheet["K" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["new"]
                sheet["L" + str(int(rowsearch) + int(sheetcellcount))] = response["deaths"]["1M_pop"]
                sheet["M" + str(int(rowsearch) + int(sheetcellcount))] = response["tests"]["total"]
                sheet["N" + str(int(rowsearch) + int(sheetcellcount))] = response["tests"]["1M_pop"]
                sheetcellcount += 1

        workbook.save(name)
        print("Pesquisa salva na planilha" + str(name) + ".")
